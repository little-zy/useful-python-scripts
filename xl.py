import openpyxl as xl
wb1 = xl.load_workbook('1.xlsx')
wb2 = xl.load_workbook('2.xlsx')
ws1 = wb1.active
ws2 = wb2.active

for row in range(1, ws1.max_row):
    for row2 in range(1, ws2.max_row):
        if(ws1.cell(row,1).value==ws2.cell(row2,3).value):
            ws1.cell(row, 15).value = ws2.cell(row2, 1).value
            break
wb1.save('3.xlsx')
