import matplotlib.pyplot as plt
import openpyxl as xl
wb = xl.load_workbook('D:\\VS_Code_Files\\Coding\\Python\\10.xlsx')
ws = wb.active
B = []
I1 = []
I2 = []
I3 = []
I4 = []
for row in range(2, ws.max_row):
    B.append(ws.cell(row, 2).value)
    I1.append(ws.cell(row, 3).value)
    I2.append(ws.cell(row, 4).value)
    I3.append(ws.cell(row, 5).value)
    I4.append(ws.cell(row, 6).value)
B2 = []
for i in B:
    B2.append(-i)
plt.plot(B2, I1, '-.', label='I1')
plt.plot(B2, I1, 'o')
plt.plot(B, I2, '-.', label='I2')
plt.plot(B, I2, 'o')
plt.plot(B2, I3, '-', label='I3')
plt.plot(B2, I3, 'o')
plt.plot(B, I4, '-', label='I4')
plt.plot(B, I4, 'o')
plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
plt.rcParams['axes.unicode_minus']=False # 用来正常显示负号
plt.xlabel('磁感应强度B/Gauss')
plt.ylabel('输出电压U_out/mV')
plt.legend()
plt.show()
